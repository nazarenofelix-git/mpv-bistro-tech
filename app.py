import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import yaml
import bcrypt
from datetime import datetime

ARQUIVO = "MPV-Controle_de_Caixa.xlsx"

# ---------------- AUTH ----------------

def load_users():
    with open("users.yaml") as f:
        return yaml.safe_load(f)

def login(user, password):
    users = load_users()["credentials"]["usernames"]

    if user in users:
        if bcrypt.checkpw(password.encode(), users[user]["password"].encode()):
            return True
    return False

if "auth" not in st.session_state:
    st.session_state.auth = False

if not st.session_state.auth:
    st.title("🔐 MPV BISTRÔ TECH")

    u = st.text_input("Usuário")
    p = st.text_input("Senha", type="password")

    if st.button("Entrar"):
        if login(u, p):
            st.session_state.auth = True
            st.session_state.user = u
            st.rerun()
        else:
            st.error("Login inválido")

    st.stop()

# ---------------- MENU ----------------

st.sidebar.title("MPV Bistrô Tech")
menu = st.sidebar.radio("Menu", ["Novo", "Registros"])

# ---------------- EXCEL ----------------

def load_df():
    return pd.read_excel(ARQUIVO, sheet_name="EVENTOS", header=2)

def insert_row(data):
    wb = load_workbook(ARQUIVO)
    ws = wb["EVENTOS"]

    ws.insert_rows(3)

    for col, val in enumerate(data, start=1):
        ws.cell(row=3, column=col, value=val)

    wb.save(ARQUIVO)

# ---------------- FORM ----------------

if menu == "Novo":

    st.title("Novo Lançamento")

    col1, col2 = st.columns(2)

    with col1:
        data = st.date_input("Data")
        desc = st.text_input("Descrição")
        valor = st.number_input("Valor")

    with col2:
        tipo = st.selectbox("Tipo", ["ENTRADA", "SAÍDA"])
        conta = st.text_input("Conta")
        centro = st.text_input("Centro de Custo")

    doc = st.text_input("Documento")
    meio = st.selectbox("Meio", ["PIX", "Cartão", "Dinheiro"])
    obs = st.text_area("Observação")

    if st.button("Salvar"):

        agora = datetime.now()

        linha = [
            data,
            desc,
            valor,
            tipo,
            conta,
            "",  # origem
            centro,
            doc,
            meio,
            "", "", "", "", "", "",
            data,  # vencimento
            "", "",
            "", "", "",
            obs,
            data,
            data.month,
            data.year,
            valor,
            "", "", "", "", "", "", "", "", "", "",
            "", "", valor,
            0,
            st.session_state.user,
            "APP",
            "", "", "", "",
            data.day,
            data.month,
            data.year,
            agora,
            data.day,
            data.month,
            data.year
        ]

        insert_row(linha)

        st.success("✅ Salvo!")

# ---------------- VIEW ----------------

if menu == "Registros":

    st.title("Registros")

    df = load_df()

    busca = st.text_input("Buscar")

    if busca:
        df = df[df.astype(str).apply(lambda x: x.str.contains(busca, case=False)).any(axis=1)]

    st.dataframe(df, use_container_width=True)
